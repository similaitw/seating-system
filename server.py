"""
Local dev server for the seating system.

Features:
- Serves static files under /web/ and /art/
- Python auto-arrange API backed by utils/auto_arrange.py
- Optional export endpoints (PDF/XLSX) when dependencies are installed

Run:
  python server.py --port 8000
Then open:
  http://localhost:8000/web/
"""

from __future__ import annotations

import argparse
import json
import mimetypes
import os
import traceback
from dataclasses import asdict, is_dataclass
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import unquote, urlparse

from models.classroom import Classroom
from models.seating import SeatingArrangement
from models.student import Student
from utils.auto_arrange import AutoArrange
from utils import rules as rules_module
from utils import optimize as optimize_module


ROOT = Path(__file__).resolve().parent
WEB_DIR = ROOT / "web"
ART_DIR = ROOT / "art"


def _json_default(obj: Any):
    if is_dataclass(obj):
        return asdict(obj)
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")


def _read_json(handler: BaseHTTPRequestHandler) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    try:
        length = int(handler.headers.get("Content-Length") or "0")
    except ValueError:
        return None, "Invalid Content-Length"
    if length <= 0:
        return None, "Empty request body"
    raw = handler.rfile.read(length)
    try:
        return json.loads(raw.decode("utf-8")), None
    except Exception:
        return None, "Invalid JSON"


def _coerce_gender(value: Any) -> str:
    v = str(value or "").strip().lower()
    if v in {"男", "m", "male", "man", "boy"}:
        return "男"
    if v in {"女", "f", "female", "woman", "girl"}:
        return "女"
    return "男"


def _coerce_students(raw_students: Any) -> List[Student]:
    students: List[Student] = []
    seen_ids: set[str] = set()
    for idx, raw in enumerate(raw_students or []):
        if not isinstance(raw, dict):
            continue

        seat_number = raw.get("seat_number", raw.get("seatNumber", raw.get("seatnumber")))
        try:
            seat_number_int = int(seat_number) if seat_number is not None and seat_number != "" else (idx + 1)
        except Exception:
            seat_number_int = idx + 1

        student_id = str(raw.get("id") or raw.get("student_id") or raw.get("studentId") or "").strip()
        if not student_id:
            student_id = f"S{seat_number_int:02d}"
        while student_id in seen_ids:
            student_id = student_id + "_" + os.urandom(2).hex()
        seen_ids.add(student_id)

        fixed = raw.get("fixed_position", raw.get("fixedPosition"))
        fixed_position = None
        if isinstance(fixed, (list, tuple)) and len(fixed) == 2:
            try:
                fixed_position = (int(fixed[0]), int(fixed[1]))
            except Exception:
                fixed_position = None

        students.append(
            Student(
                id=student_id,
                seat_number=seat_number_int,
                name=str(raw.get("name") or raw.get("姓名") or ""),
                gender=_coerce_gender(raw.get("gender") or raw.get("性別") or raw.get("sex")),
                height=(int(raw["height"]) if raw.get("height") not in (None, "") else None),
                vision_left=float(raw.get("vision_left", raw.get("visionLeft", 1.0)) or 1.0),
                vision_right=float(raw.get("vision_right", raw.get("visionRight", 1.0)) or 1.0),
                need_front_seat=bool(raw.get("need_front_seat", raw.get("needFrontSeat", False))),
                need_aisle_seat=bool(raw.get("need_aisle_seat", raw.get("needAisleSeat", False))),
                need_near_teacher=bool(raw.get("need_near_teacher", raw.get("needNearTeacher", False))),
                fixed_position=fixed_position,
                notes=str(raw.get("notes") or ""),
            )
        )

    return students


def _coerce_classroom(raw: Dict[str, Any]) -> Classroom:
    # Classroom.from_dict expects fields present; provide defaults.
    payload = {
        "name": raw.get("name", "Classroom"),
        "rows": int(raw.get("rows", 5)),
        "cols": int(raw.get("cols", 6)),
        "teacher_desk_position": raw.get("teacher_desk_position", raw.get("teacherDeskPosition", "front")),
        "special_seats": raw.get("special_seats", raw.get("specialSeats", [])) or [],
        "empty_seats": raw.get("empty_seats", raw.get("emptySeats", [])) or [],
        "orientation": raw.get("orientation", "front"),
    }
    return Classroom.from_dict(payload)


def _normalize_seats_matrix(seats: Any, rows: int, cols: int) -> List[List[Optional[str]]]:
    matrix: List[List[Optional[str]]] = [[None for _ in range(cols)] for _ in range(rows)]
    if not isinstance(seats, list):
        return matrix
    for r in range(min(rows, len(seats))):
        row = seats[r]
        if not isinstance(row, list):
            continue
        for c in range(min(cols, len(row))):
            v = row[c]
            matrix[r][c] = str(v) if v not in (None, "") else None
    return matrix


def _coerce_arrangement(raw: Dict[str, Any], classroom: Classroom) -> SeatingArrangement:
    arrangement_id = str(raw.get("id") or "arr_" + os.urandom(4).hex())
    payload = {
        "id": arrangement_id,
        "name": raw.get("name", "Seating Arrangement"),
        "classroom_id": raw.get("classroom_id", raw.get("classroomId", "classroom_default")),
        "created_at": raw.get("created_at", raw.get("createdAt")),
        "seats": _normalize_seats_matrix(raw.get("seats"), classroom.rows, classroom.cols),
        "locked_seats": raw.get("locked_seats", raw.get("lockedSeats", [])) or [],
    }

    # Ensure created_at exists
    if not payload["created_at"]:
        payload.pop("created_at", None)

    arrangement = SeatingArrangement.from_dict(payload)

    # Remove locked seats that are outside bounds or overlap empty seats.
    normalized_locked: List[Tuple[int, int]] = []
    empty = set((int(r), int(c)) for r, c in classroom.empty_seats)
    for r, c in arrangement.locked_seats:
        try:
            rr, cc = int(r), int(c)
        except Exception:
            continue
        if not (0 <= rr < classroom.rows and 0 <= cc < classroom.cols):
            continue
        if (rr, cc) in empty:
            continue
        normalized_locked.append((rr, cc))
    arrangement.locked_seats = normalized_locked
    return arrangement


def _run_auto_arrange(mode: str, classroom: Classroom, students: List[Student], arrangement: SeatingArrangement) -> SeatingArrangement:
    if mode == "seat_number":
        return AutoArrange.by_seat_number(students, classroom, arrangement)
    if mode == "alternating_gender":
        return AutoArrange.alternating_gender(students, classroom, arrangement)
    if mode == "by_height":
        return AutoArrange.by_height(students, classroom, arrangement)
    if mode == "by_vision":
        return AutoArrange.by_vision(students, classroom, arrangement)
    if mode == "random":
        return AutoArrange.random_arrange(students, classroom, arrangement)
    raise ValueError(f"Unsupported mode: {mode}")


def _guess_mime(path: Path) -> str:
    mime, _ = mimetypes.guess_type(str(path))
    return mime or "application/octet-stream"


def _is_within(path: Path, base: Path) -> bool:
    try:
        path.resolve().relative_to(base.resolve())
        return True
    except Exception:
        return False


def _read_static(base: Path, url_path: str) -> Tuple[Optional[Path], Optional[str]]:
    rel = url_path.lstrip("/")
    if not rel:
        return None, "Empty path"
    # Prevent traversal
    if ".." in rel.replace("\\", "/").split("/"):
        return None, "Invalid path"
    file_path = (base / rel).resolve()
    if not _is_within(file_path, base):
        return None, "Invalid path"
    if file_path.is_dir():
        file_path = file_path / "index.html"
    if not file_path.exists():
        return None, "Not found"
    if not file_path.is_file():
        return None, "Not a file"
    return file_path, None


class Handler(BaseHTTPRequestHandler):
    server_version = "SeatingSystemServer/1.0"

    def log_message(self, fmt: str, *args):
        # Less noisy
        return

    def _send_json(self, status: int, payload: Dict[str, Any]):
        raw = json.dumps(payload, ensure_ascii=False, indent=2, default=_json_default).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def _send_bytes(self, status: int, content: bytes, content_type: str, filename: Optional[str] = None):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store")
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def _send_text(self, status: int, text: str, content_type: str = "text/plain; charset=utf-8"):
        data = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _serve_file(self, file_path: Path):
        try:
            content = file_path.read_bytes()
        except Exception:
            self._send_text(HTTPStatus.INTERNAL_SERVER_ERROR, "Failed to read file")
            return
        self._send_bytes(HTTPStatus.OK, content, _guess_mime(file_path))

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/api/health":
            exports = {
                "pdf": _pdf_available(),
                "xlsx": _xlsx_available(),
            }
            self._send_json(HTTPStatus.OK, {"ok": True, "version": 1, "exports": exports})
            return

        if path == "/" or path == "/web" or path == "/web/":
            file_path = WEB_DIR / "index.html"
            if file_path.exists():
                self._serve_file(file_path)
            else:
                self._send_text(HTTPStatus.NOT_FOUND, "Missing web/index.html")
            return

        if path.startswith("/web/"):
            file_path, err = _read_static(WEB_DIR, path[len("/web/"):])
            if err:
                self._send_text(HTTPStatus.NOT_FOUND, err)
                return
            self._serve_file(file_path)
            return

        if path.startswith("/art/"):
            file_path, err = _read_static(ART_DIR, path[len("/art/"):])
            if err:
                self._send_text(HTTPStatus.NOT_FOUND, err)
                return
            self._serve_file(file_path)
            return

        self._send_text(HTTPStatus.NOT_FOUND, "Not found")

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/api/auto-arrange":
            body, err = _read_json(self)
            if err:
                self._send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": err})
                return
            try:
                mode = str(body.get("mode") or "")
                payload = body.get("payload") or body
                classroom = _coerce_classroom(payload.get("classroom") or {})
                students = _coerce_students(payload.get("students") or [])
                arrangement = _coerce_arrangement(payload.get("arrangement") or {}, classroom)
                rules_dict = payload.get("rules") or {}

                if mode == "optimize":
                    # 先用 seat_number 取得一個合法的初始解
                    seed_mode = str(body.get("seed_mode") or "seat_number")
                    if seed_mode not in {"seat_number", "alternating_gender", "by_height", "by_vision", "random"}:
                        seed_mode = "seat_number"
                    seeded = _run_auto_arrange(seed_mode, classroom, students, arrangement)

                    time_budget = float(body.get("time_budget") or 3.0)
                    time_budget = max(0.5, min(time_budget, 15.0))

                    result = optimize_module.optimize(
                        classroom=classroom.to_dict(),
                        students=[s.to_dict() for s in students],
                        arrangement=seeded.to_dict(),
                        rules=rules_dict,
                        time_budget=time_budget,
                    )
                    self._send_json(HTTPStatus.OK, {
                        "ok": True,
                        "arrangement": result["arrangement"],
                        "violations": result["violations"],
                        "score": result["score"],
                        "max_score": result["max_score"],
                        "by_type": result["by_type"],
                        "stats": {
                            "iterations": result["iterations"],
                            "accepted": result["accepted"],
                            "improved": result["improved"],
                            "elapsed": result["elapsed"],
                            "seed_mode": seed_mode,
                        },
                    })
                    return

                result = _run_auto_arrange(mode, classroom, students, arrangement)

                # 以排座結果做規則評分（與前端規則對齊）
                scored = rules_module.evaluate(
                    classroom=classroom.to_dict(),
                    students=[s.to_dict() for s in students],
                    arrangement=result.to_dict(),
                    rules=rules_dict,
                )
                self._send_json(HTTPStatus.OK, {
                    "ok": True,
                    "arrangement": result.to_dict(),
                    "violations": scored["violations"],
                    "score": scored["score"],
                    "max_score": scored["max_score"],
                    "by_type": scored["by_type"],
                })
            except Exception as ex:
                self._send_json(
                    HTTPStatus.BAD_REQUEST,
                    {"ok": False, "error": str(ex), "trace": traceback.format_exc(limit=4)},
                )
            return

        if path == "/api/evaluate":
            body, err = _read_json(self)
            if err:
                self._send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": err})
                return
            try:
                payload = body.get("payload") or body
                classroom = _coerce_classroom(payload.get("classroom") or {})
                students = _coerce_students(payload.get("students") or [])
                arrangement = _coerce_arrangement(payload.get("arrangement") or {}, classroom)
                scored = rules_module.evaluate(
                    classroom=classroom.to_dict(),
                    students=[s.to_dict() for s in students],
                    arrangement=arrangement.to_dict(),
                    rules=payload.get("rules") or {},
                )
                self._send_json(HTTPStatus.OK, {
                    "ok": True,
                    "violations": scored["violations"],
                    "score": scored["score"],
                    "max_score": scored["max_score"],
                    "by_type": scored["by_type"],
                })
            except Exception as ex:
                self._send_json(
                    HTTPStatus.BAD_REQUEST,
                    {"ok": False, "error": str(ex), "trace": traceback.format_exc(limit=4)},
                )
            return

        if path == "/api/export/xlsx":
            body, err = _read_json(self)
            if err:
                self._send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": err})
                return
            if not _xlsx_available():
                self._send_json(
                    HTTPStatus.NOT_IMPLEMENTED,
                    {"ok": False, "error": "openpyxl not installed. Run: pip install -r requirements.txt"},
                )
                return
            try:
                payload = body.get("payload") or body
                filename = body.get("filename") or "seating.xlsx"
                content = _export_xlsx(payload)
                self._send_bytes(HTTPStatus.OK, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)
            except Exception as ex:
                self._send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(ex), "trace": traceback.format_exc(limit=4)})
            return

        if path == "/api/export/pdf":
            body, err = _read_json(self)
            if err:
                self._send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": err})
                return
            if not _pdf_available():
                self._send_json(
                    HTTPStatus.NOT_IMPLEMENTED,
                    {"ok": False, "error": "reportlab not installed. Run: pip install -r requirements.txt"},
                )
                return
            try:
                payload = body.get("payload") or body
                filename = body.get("filename") or "seating.pdf"
                content = _export_pdf(payload)
                self._send_bytes(HTTPStatus.OK, content, "application/pdf", filename=filename)
            except Exception as ex:
                self._send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(ex), "trace": traceback.format_exc(limit=4)})
            return

        self._send_text(HTTPStatus.NOT_FOUND, "Not found")


def _xlsx_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def _pdf_available() -> bool:
    try:
        import reportlab  # noqa: F401
        return True
    except Exception:
        return False


def _build_table(payload: Dict[str, Any]) -> Tuple[Classroom, List[Student], SeatingArrangement, List[List[str]]]:
    classroom = _coerce_classroom(payload.get("classroom") or {})
    students = _coerce_students(payload.get("students") or [])
    arrangement = _coerce_arrangement(payload.get("arrangement") or {}, classroom)

    student_by_id = {s.id: s for s in students}
    empty = set((int(r), int(c)) for r, c in classroom.empty_seats)
    locked = set((int(r), int(c)) for r, c in arrangement.locked_seats)

    grid: List[List[str]] = []
    for r in range(classroom.rows):
        row: List[str] = []
        for c in range(classroom.cols):
            if (r, c) in empty:
                row.append("空位")
                continue
            sid = arrangement.get_student_at(r, c)
            if not sid:
                row.append("")
                continue
            s = student_by_id.get(sid)
            label = f"{s.seat_number:02d} {s.name}" if s else sid
            if (r, c) in locked:
                label = "🔒 " + label
            row.append(label)
        grid.append(row)

    return classroom, students, arrangement, grid


def _export_xlsx(payload: Dict[str, Any]) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    classroom, students, arrangement, grid = _build_table(payload)
    empty = set((int(r), int(c)) for r, c in classroom.empty_seats)
    locked = set((int(r), int(c)) for r, c in arrangement.locked_seats)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seating"

    title = payload.get("title") or arrangement.name or "Seating"
    ws["A1"] = str(title)
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=classroom.cols + 1)

    ws["A2"] = f"Classroom: {classroom.name}"
    ws["A3"] = f"Desk: {classroom.teacher_desk_position}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=classroom.cols + 1)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=classroom.cols + 1)

    header_row = 5
    ws.cell(row=header_row, column=1, value="")
    for c in range(classroom.cols):
        ws.cell(row=header_row, column=2 + c, value=f"C{c + 1}")

    thin = Side(style="thin", color="B0AEA5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_empty = PatternFill("solid", fgColor="E8E6DC")
    fill_locked = PatternFill("solid", fgColor="DDEAF7")

    for r in range(classroom.rows):
        ws.cell(row=header_row + 1 + r, column=1, value=f"R{r + 1}")
        ws.row_dimensions[header_row + 1 + r].height = 32
        for c in range(classroom.cols):
            cell = ws.cell(row=header_row + 1 + r, column=2 + c, value=grid[r][c])
            cell.alignment = align
            cell.border = border
            if (r, c) in empty:
                cell.fill = fill_empty
            elif (r, c) in locked:
                cell.fill = fill_locked

    # Column widths
    ws.column_dimensions["A"].width = 6
    for c in range(classroom.cols):
        col_letter = openpyxl.utils.get_column_letter(2 + c)
        ws.column_dimensions[col_letter].width = 18

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _export_pdf(payload: Dict[str, Any]) -> bytes:
    from xml.sax.saxutils import escape

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    classroom, students, arrangement, grid = _build_table(payload)

    # Best-effort CJK font registration (Windows fonts).
    font_name = "Helvetica"
    for candidate, sub in [
        (r"C:\Windows\Fonts\msjh.ttc", 0),
        (r"C:\Windows\Fonts\msjhl.ttc", 0),
        (r"C:\Windows\Fonts\mingliu.ttc", 0),
        (r"C:\Windows\Fonts\simsun.ttc", 0),
        (r"C:\Windows\Fonts\simhei.ttf", 0),
    ]:
        try:
            if Path(candidate).exists():
                pdfmetrics.registerFont(TTFont("CJK", candidate, subfontIndex=sub))
                font_name = "CJK"
                break
        except Exception:
            continue

    styles = getSampleStyleSheet()
    base = ParagraphStyle(
        "Base",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        leading=11,
    )
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=14,
        leading=16,
        spaceAfter=6,
    )

    # Build table data with headers
    data: List[List[Any]] = []
    header = [""] + [f"C{c + 1}" for c in range(classroom.cols)]
    data.append([Paragraph(escape(h), base) for h in header])
    for r in range(classroom.rows):
        row = [Paragraph(escape(f"R{r + 1}"), base)]
        for c in range(classroom.cols):
            row.append(Paragraph(escape(grid[r][c] or ""), base))
        data.append(row)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=str(arrangement.name or "Seating"),
    )

    elements: List[Any] = []
    elements.append(Paragraph(escape(str(payload.get("title") or arrangement.name or "Seating"))), title_style)
    elements.append(Paragraph(escape(f"Classroom: {classroom.name} · Desk: {classroom.teacher_desk_position}"), base))
    elements.append(Spacer(1, 6 * mm))

    table = Table(data, repeatRows=1)
    table_style = TableStyle(
        [
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B0AEA5")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E6DC")),
        ]
    )

    # Highlight empty/locked seats
    empty = set((int(r), int(c)) for r, c in classroom.empty_seats)
    locked = set((int(r), int(c)) for r, c in arrangement.locked_seats)
    for r in range(classroom.rows):
        for c in range(classroom.cols):
            if (r, c) in empty:
                table_style.add("BACKGROUND", (1 + c, 1 + r), (1 + c, 1 + r), colors.HexColor("#F0EEE6"))
            elif (r, c) in locked:
                table_style.add("BACKGROUND", (1 + c, 1 + r), (1 + c, 1 + r), colors.HexColor("#DDEAF7"))

    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()


def main():
    parser = argparse.ArgumentParser(description="Seating system local server")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8000)
    args = parser.parse_args()

    if not WEB_DIR.exists():
        raise SystemExit("Missing web/ directory")

    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(f"Serving on http://{args.host}:{args.port}/web/")
    print("API: GET /api/health · POST /api/auto-arrange · POST /api/evaluate · POST /api/export/pdf · POST /api/export/xlsx")
    server.serve_forever()


if __name__ == "__main__":
    main()

