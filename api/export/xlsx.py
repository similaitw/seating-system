import traceback
from http.server import BaseHTTPRequestHandler
from io import BytesIO
from typing import Any, Dict

from api._lib.common import (
    build_table,
    read_json_body,
    send_bytes,
    send_json,
    xlsx_available,
)


def _export_xlsx(payload: Dict[str, Any]) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    classroom, students, arrangement, grid = build_table(payload)
    empty = set((int(r), int(c)) for r, c in classroom.empty_seats)
    locked = set((int(r), int(c)) for r, c in arrangement.locked_seats)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seating"

    title = payload.get("title") or arrangement.name or "Seating"
    ws["A1"] = str(title)
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=classroom.cols + 1)

    ws["A2"] = f"Classroom: {classroom.name}"
    ws["A3"] = f"Desk: {classroom.teacher_desk_position}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=classroom.cols + 1)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=classroom.cols + 1)

    header_row = 5
    ws.cell(row=header_row, column=1, value="")
    for c in range(classroom.cols):
        ws.cell(row=header_row, column=2 + c, value=f"C{c + 1}")

    thin = Side(style="thin", color="B0AEA5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_empty = PatternFill("solid", fgColor="E8E6DC")
    fill_locked = PatternFill("solid", fgColor="DDEAF7")

    for r in range(classroom.rows):
        ws.cell(row=header_row + 1 + r, column=1, value=f"R{r + 1}")
        ws.row_dimensions[header_row + 1 + r].height = 32
        for c in range(classroom.cols):
            cell = ws.cell(row=header_row + 1 + r, column=2 + c, value=grid[r][c])
            cell.alignment = align
            cell.border = border
            if (r, c) in empty:
                cell.fill = fill_empty
            elif (r, c) in locked:
                cell.fill = fill_locked

    ws.column_dimensions["A"].width = 6
    for c in range(classroom.cols):
        col_letter = openpyxl.utils.get_column_letter(2 + c)
        ws.column_dimensions[col_letter].width = 18

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        body, err = read_json_body(self)
        if err:
            send_json(self, 400, {"ok": False, "error": err})
            return
        if not xlsx_available():
            send_json(self, 501, {"ok": False, "error": "openpyxl not installed"})
            return
        try:
            payload = body.get("payload") or body
            filename = body.get("filename") or "seating.xlsx"
            content = _export_xlsx(payload)
            send_bytes(
                self, 200, content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=filename,
            )
        except Exception as ex:
            send_json(self, 400, {
                "ok": False,
                "error": str(ex),
                "trace": traceback.format_exc(limit=4),
            })
